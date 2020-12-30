#!/usr/bin/env python
# coding: utf-8

# # find out the file data 
# 

# In[1]:


import  pandas  as pd
from pandas import DataFrame
from pandas import Series,concat


# In[2]:


from openpyxl import load_workbook


# In[3]:


import os
import xlrd


# In[4]:


import os

dirname = r"银川二污普数据/最新数据/二污普导出数据/普查报表/银川市工业源报表-2743"
filename = ".xlsx"

result = []
def search(dirname=dirname, filename=""):
    for item in os.listdir(dirname):
        item_path = os.path.join(dirname, item)
        if os.path.isdir(item_path):
            search(item_path, filename)
        elif os.path.isfile(item_path):
            if filename in item:
                global result
                result.append(item_path)
                #print(item_path+";")


# In[5]:


search(dirname, filename)


# # final work !!!!

# In[6]:


a_old=result[:30]


# In[7]:



class Fib():                  #定义类Fib
    def __init__(self):
        self.changed={}         #给类一个字典属性，方便后续的存储
    def __getitem__(self, key): #定性__getitem__函数，key为类Fib的键
        return self.changed[key] #当按照键取值时，返回的值为changed[key]的值
    def __setitem__(self,key,value):
        self.changed[key]=value 

s=Fib()


# In[8]:


c=[]
#检查表格是否完整，将不完整的存入list
    
for i in range (len(a_old)):
        
    s[i]=load_workbook(filename=a_old[i])
    if len(s[i].sheetnames)<=1:
        c.append(i)
    elif s[i].sheetnames[1]!='T_BAS_G101_2':
        c.append(i)


# In[9]:


a=[e for i ,e in enumerate (a_old) if i not in c]


# In[10]:


#get the data
for i, e in enumerate (a):
    s[i]=xlrd.open_workbook(filename=a[i])


# In[11]:


title=s[1].sheet_by_name(u'T_BAS_G101_2').row_values(6)


# In[12]:


i=0
z_1=[]
while i != len(a):
    t=s[i].sheet_by_name(u'T_BAS_G101_2')
    z_1.append([])
    d=7
    while t.row_values(d)[1]!='单位负责人： ':
        z_1[i].append(t.row_values(d))
        d+=1
    i+=1


# In[13]:


df = pd.DataFrame()
i=0
while i != len(z_1):
    df=df.append(z_1[i])
    i+=1


# In[14]:


df.columns=title


# In[15]:


tst_1=z_1


# In[16]:



class Fib():                  #定义类Fib
    def __init__(self):
        self.changed={}         #给类一个字典属性，方便后续的存储
    def __getitem__(self, key): #定性__getitem__函数，key为类Fib的键
        return self.changed[key] #当按照键取值时，返回的值为changed[key]的值
    def __setitem__(self,key,value):
        self.changed[key]=value 

s_c=Fib()


# In[17]:


def fun (c,y):
    for i in range (len(a)):
        s_c[i]=load_workbook(filename=a[i])
        c.append(i)
        c[i]=s_c[i][s_c[i].sheetnames[1]][y].value
    return('finish')
        


# In[18]:


普查小区代码=[]


# In[19]:


fun (普查小区代码,'D2')


# In[20]:


统一社会信用代码=[]


# In[21]:


fun (统一社会信用代码,'D3')


# In[22]:


组织机构代码=[]


# In[23]:


fun (组织机构代码,'D4')


# In[24]:


单位详细名称  =[]


# In[25]:


fun (单位详细名称,'D5')


# In[26]:


for i in range (len(tst_1)):
    for e in range (len(tst_1[i])):
        tst_1[i][e][0]=单位详细名称[i]
        tst_1[i][e][1]=普查小区代码[i]
        tst_1[i][e][2]=统一社会信用代码[i]
        tst_1[i][e][3]=组织机构代码[i]                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
        
    #tst_1[i][0][0]=普查小区代码[i]


# In[27]:


cdf_t=pd.DataFrame()
i=0
while i != len(z_1):
    cdf_t=cdf_t.append(tst_1[i])
    i+=1


# In[28]:


cdf_t.columns=['单位详细名称','普查小区代码','统一社会信用代码','组织机构代码',1,2,3,4,5]


# In[29]:


final_df= pd.concat([df,cdf_t],axis=1)


# In[30]:


final_df=final_df.drop([1,2,3,4,5],axis=1)


# In[31]:


final_df.to_csv('data/T_BAS_G101_2.csv',encoding='utf-8-sig')

