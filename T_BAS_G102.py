#!/usr/bin/env python
# coding: utf-8

# # find out the file data 
# 

# In[1]:


import  pandas  as pd
from pandas import DataFrame
from pandas import Series,concat


# In[2]:


from openpyxl import load_workbook


# In[3]:


import os
import xlrd


# In[4]:


import os

dirname = r"E:/宁夏二污普/最新数据/二污普导出数据/普查报表/银川市工业源报表-2743"
filename = ".xlsx"

result = []
def search(dirname=dirname, filename=""):
    for item in os.listdir(dirname):
        item_path = os.path.join(dirname, item)
        if os.path.isdir(item_path):
            search(item_path, filename)
        elif os.path.isfile(item_path):
            if filename in item:
                global result
                result.append(item_path)
                #print(item_path+";")


# In[5]:


search(dirname, filename)


# # final work !!!!

# In[6]:


for i in range(len(result)):
    if result[i]=='E:/宁夏二污普/最新数据/二污普导出数据/普查报表/银川市工业源报表-2743\金凤区-145\运行、停产、关闭-104\宁夏报业传媒印刷有限公司.xlsx':
        print(i)


# In[7]:


a_old=result[2665:5675]


# In[8]:



class Fib():                  #定义类Fib
    def __init__(self):
        self.changed={}         #给类一个字典属性，方便后续的存储
    def __getitem__(self, key): #定性__getitem__函数，key为类Fib的键
        return self.changed[key] #当按照键取值时，返回的值为changed[key]的值
    def __setitem__(self,key,value):
        self.changed[key]=value 

s=Fib()


# In[9]:


a=[]
#检查表格是否完整，将不完整的存入list
    
for i in range (len(a_old)):    
    s[i]=load_workbook(filename=a_old[i])
    if 'T_BAS_G102' in s[i].sheetnames:
        a.append(a_old[i])


# In[10]:


#get the data
for i, e in enumerate (a):
    s[i]=xlrd.open_workbook(filename=a[i])


# In[11]:


i=0
z_1=[]
xt=[]
while i != len(a):
    t=s[i].sheet_by_name(u'T_BAS_G102')
    z_1.append([])
    z_1[i].append(t.row_values(15))
    z_1[i].append(t.row_values(16))
    d=19
    while d !=31:
        z_1[i].append(t.row_values(d))
        d+=1
    i+=1


# In[12]:


x_len=[]
for i in range(len(z_1)):
    x_len.append(len(z_1[i][0]))


# In[13]:


#let the length be the same 
for i in range(len(z_1)):
    for row in range (len(z_1[i])):
        #print(len(z_1[i][row]))
        if len(z_1[i][row]) < max(x_len):
            z_1[i][row].extend(['']* (max(x_len)-len(z_1[i][row])))


# In[14]:


xz=z_1[27]


# In[15]:


def fun(xz):
    xt=[]
    for i in range(len (xz[1])):
        xt.append([])
    x=0
    while x != len(xz):
        e=0
        while e!= len(xz[1]):
            xt[e].append(xz[x][e])
            e+=1
        x+=1
    return(xt)


# In[16]:


xy=[]
for i in range (len(z_1)):
    xy.append(fun(z_1[i]))


# In[17]:


df_1 = pd.DataFrame()
i=0
while i != len(xy):
    df_1=df_1.append(xy[i])
    i+=1


# In[18]:


df_1.to_csv('E:/cleandata/T_BAS_xy_test.csv',encoding='utf-8-sig')


# In[19]:


tst_1=xy


# In[20]:


普查小区代码=[]
统一社会信用代码=[]
组织机构代码=[]
单位详细名称  =[]
取水量 =[]
城市自来水=[]
自备水=[]
水利工程供水=[]
其他工业企业供水=[]
for i in range(len(a)):
    普查小区代码.append(s[i].sheet_by_name(u'T_BAS_G102').cell(1,3).value)
    统一社会信用代码.append(s[i].sheet_by_name(u'T_BAS_G102').cell(2,3).value)
    组织机构代码.append(s[i].sheet_by_name(u'T_BAS_G102').cell(3,3).value)
    单位详细名称.append(s[i].sheet_by_name(u'T_BAS_G102').cell(4,3).value)
    取水量.append(s[i].sheet_by_name(u'T_BAS_G102').cell(9,5).value)
    城市自来水.append(s[i].sheet_by_name(u'T_BAS_G102').cell(10,5).value)
    自备水.append(s[i].sheet_by_name(u'T_BAS_G102').cell(11,5).value)
    水利工程供水.append(s[i].sheet_by_name(u'T_BAS_G102').cell(12,5).value)
    其他工业企业供水.append(s[i].sheet_by_name(u'T_BAS_G102').cell(13,5).value)


# In[21]:


for i in range (len(tst_1)):
    for e in range (len(tst_1[i])):
        tst_1[i][e][0]=单位详细名称[i]
        tst_1[i][e][1]=普查小区代码[i]
        tst_1[i][e][2]=统一社会信用代码[i]
        tst_1[i][e][3]=组织机构代码[i]                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
        tst_1[i][e][4]=取水量[i]
        tst_1[i][e][5]=城市自来水[i]
        tst_1[i][e][6]=自备水[i]
        tst_1[i][e][7]=水利工程供水[i]
        tst_1[i][e][8]=其他工业企业供水[i]
    #tst_1[i][0][0]=普查小区代码[i]


# In[22]:


cdf_t=pd.DataFrame()
i=0
while i != len(z_1):
    cdf_t=cdf_t.append(tst_1[i])
    i+=1


# In[24]:


cdf_t.columns=['单位详细名称','普查小区代码','统一社会信用代码','组织机构代码','取水量','城市自来水','自备水','水利工程供水','其他工业企业供水','x_1','x_2','x_3','x_4','x_5']


# In[25]:


final_df= pd.concat([df_1,cdf_t],axis=1)


# In[26]:


final_df=final_df.drop(['x_1','x_2','x_3','x_4','x_5'],axis=1)


# In[27]:


final_df.to_csv('E:/cleandata/T_BAS_G102.csv',encoding='utf-8-sig')


# In[28]:


data=pd.read_csv('E:/cleandata/T_BAS_G102.csv',encoding='utf-8-sig')


# In[29]:


data=data.dropna(subset=['0'])


# In[30]:


data=data.reset_index(drop=True)


# In[31]:


for i,e in enumerate (data['0']):
    if e =='甲':
        data=data.drop(labels=[i],axis=0)
    elif e =='乙':
        data=data.drop(labels=[i],axis=0)
    elif e=='丙':
        data=data.drop(labels=[i],axis=0)


# In[32]:


data.to_csv('E:/cleandata/T_BAS_G102.csv',encoding='utf-8-sig')


# In[ ]:




