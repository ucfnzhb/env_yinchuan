#!/usr/bin/env python
# coding: utf-8

# # find out the file data 
# 

# In[1]:


import  pandas  as pd
from pandas import DataFrame
from pandas import Series,concat


# In[2]:


from openpyxl import load_workbook


# In[3]:


import os
import xlrd


# In[4]:


import os

dirname = r"宁夏二污普/最新数据/二污普导出数据/普查报表/银川市工业源报表-2743"
filename = ".xlsx"

result = []
def search(dirname=dirname, filename=""):
    for item in os.listdir(dirname):
        item_path = os.path.join(dirname, item)
        if os.path.isdir(item_path):
            search(item_path, filename)
        elif os.path.isfile(item_path):
            if filename in item:
                global result
                result.append(item_path)
                #print(item_path+";")


# In[5]:


search(dirname, filename)


# # final work !!!!




# In[8]:


a_old=result


# In[9]:



class Fib():                  #定义类Fib
    def __init__(self):
        self.changed={}         #给类一个字典属性，方便后续的存储
    def __getitem__(self, key): #定性__getitem__函数，key为类Fib的键
        return self.changed[key] #当按照键取值时，返回的值为changed[key]的值
    def __setitem__(self,key,value):
        self.changed[key]=value 

s=Fib()


# In[10]:


a=[]
#检查表格是否完整，将不完整的存入list
    
for i in range (len(a_old)):
        
    s[i]=load_workbook(filename=a_old[i])
    if 'T_BAS_G101_3' in s[i].sheetnames:
        a.append(a_old[i])




# In[12]:


#get the data
for i, e in enumerate (a):
    s[i]=xlrd.open_workbook(filename=a[i])




title=s[1].sheet_by_name(u'T_BAS_G101_3').row_values(6)


# In[15]:


i=0
z_1=[]
z_2=[]
while i != len(a):
    t=s[i].sheet_by_name(u'T_BAS_G101_3')
    z_1.append([])
    z_2.append([])
    d=8
    while t.row_values(d)[2]!='二、主要能源消耗':
        z_1[i].append(t.row_values(d))
        d+=1
    x=d+1
    while t.row_values(x)[1]!='单位负责人： ':
        z_2[i].append(t.row_values(x))
        x+=1
    i+=1


# In[16]:


df = pd.DataFrame()
i=0
while i != len(z_1):
    df=df.append(z_1[i])
    i+=1


# In[17]:


df_2 = pd.DataFrame()
i=0
while i != len(z_2):
    df_2=df_2.append(z_2[i])
    i+=1


# In[18]:


df.columns=title


# In[19]:


df_2.columns=title


# In[20]:


tst_1=z_1


# In[21]:


tst_2=z_2


# In[22]:


普查小区代码=[]
统一社会信用代码=[]
组织机构代码=[]
单位详细名称  =[]
for i in range(len(a)):
    普查小区代码.append(s[i].sheet_by_name(u'T_BAS_G101_3').cell(1,3).value)
    统一社会信用代码.append(s[i].sheet_by_name(u'T_BAS_G101_3').cell(2,3).value)
    组织机构代码.append(s[i].sheet_by_name(u'T_BAS_G101_3').cell(3,3).value)
    单位详细名称.append(s[i].sheet_by_name(u'T_BAS_G101_3').cell(4,3).value)


# In[23]:


for i in range (len(tst_1)):
    for e in range (len(tst_1[i])):
        tst_1[i][e][0]=单位详细名称[i]
        tst_1[i][e][1]=普查小区代码[i]
        tst_1[i][e][2]=统一社会信用代码[i]
        tst_1[i][e][3]=组织机构代码[i]                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
        
    #tst_1[i][0][0]=普查小区代码[i]


# In[24]:


for i in range (len(tst_2)):
    for e in range (len(tst_2[i])):
        tst_2[i][e][0]=单位详细名称[i]
        tst_2[i][e][1]=普查小区代码[i]
        tst_2[i][e][2]=统一社会信用代码[i]
        tst_2[i][e][3]=组织机构代码[i]                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                 
        
    #tst_1[i][0][0]=普查小区代码[i]


# In[25]:


cdf_t=pd.DataFrame()
i=0
while i != len(z_1):
    cdf_t=cdf_t.append(tst_1[i])
    i+=1


# In[26]:


cdf_2=pd.DataFrame()
i=0
while i != len(z_2):
    cdf_2=cdf_2.append(tst_2[i])
    i+=1


# In[27]:


cdf_t.columns=['单位详细名称','普查小区代码','统一社会信用代码','组织机构代码',1,2,3,4,5]


# In[28]:


cdf_2.columns=['单位详细名称','普查小区代码','统一社会信用代码','组织机构代码',1,2,3,4,5]


# In[29]:


final_df= pd.concat([df,cdf_t],axis=1)


# In[30]:


final_df_2= pd.concat([df_2,cdf_2],axis=1)


# In[31]:


final_df=final_df.drop([1,2,3,4,5],axis=1)


# In[32]:


final_df_2=final_df_2.drop([1,2,3,4,5],axis=1)


# In[33]:


final_df.to_csv('data/T_BAS_G101_3_原辅材料.csv',encoding='utf-8-sig')


# In[34]:


final_df_2.to_csv('data/T_BAS_G101_3_能源名称.csv',encoding='utf-8-sig')

